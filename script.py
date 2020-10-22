import os, sys, csv
import openpyxl
from openpyxl.worksheet.table import Table
from copy import copy
from itertools import chain

# context of code is Akvakulturregisteret.csv from https://fiskeridir.no/Akvakultur/Registre-og-skjema/Akvakulturregisteret

# mapping row index to text 
index_values = ['TILL_NR', 'ORG.NR/PERS.NR', 'NAVN', 'ADRESSE', 'POSTNR', 'POSTSTED', 'TILDELINGSTIDSPUNKT', 'TIDSBEGRENSET', 'TILL_KOMNR', 'TILL_KOM', 'FORMÅL', 'PRODUKSJONSFORM', 'ART', 'TILL_KAP', 'TILL_ENHET', 'LOK_NR', 'LOK_NAVN', 'LOK_KOMNR', 'LOK_KOM', 'LOK_PLASS', 'VANNMILJØ', 'LOK_KAP', 'LOK_ENHET', 'UTGÅR_DATO', 'N_GEOWGS84', 'Ø_GEOWGS84']

#  row sub index for data 
ORG_NR = 1
LOK_NR = 15

   
def samlokalisert(in_filename):
  wb = openpyxl.load_workbook(in_filename)
  ws = wb.active
  locations = {}

  # table in xlsx not found in ws.tables.values(), magic?
  for i, row in enumerate(ws.iter_rows()):
    if not row[LOK_NR].value or row[LOK_NR].value in ["", '', "LOK_NR"]:
      continue
    start_row = i
    break
  
  if not start_row:
    print("failed")
    exit(1)

  for row in ws.iter_rows(min_row = start_row):    
    org = row[ORG_NR].value
    lok = row[LOK_NR].value 
    if lok not in locations:
      locations[lok] = set()
    locations[lok].add(org)

  colocations = { lok_nr for (lok_nr, v) in locations.items() if len(v) >= 2}      

  # list lokasjoner som er knyttet til samlokalisering
  ws2 = wb.create_sheet("Samlokasjoner")
  for row in ws.iter_rows():
    values = [cell.value for cell in row]
    if row[LOK_NR].value and row[LOK_NR].value.isnumeric():
      if row[LOK_NR].value in colocations:
        ws2.append(values)        
      else:
        continue
    else:
      ws2.append(values)

  get_column_letter = openpyxl.utils.cell.get_column_letter
  dims = get_column_letter(ws2.min_column) + str(start_row) + ":" + get_column_letter(ws2.max_column) + str(ws2.max_row) 
  table = Table(ref=dims, displayName='samlokasjon')
  ws2.add_table(table)


  # list org_nr som driver samlokalisering i ark 3, separert med ';'
  ws3 = wb.create_sheet("Samlokaliserte ORG-NR")
  unique_orgs = {org for org in chain.from_iterable(locations.values())}
  text = ""
  for org in unique_orgs:
    text += str(org) + ";"  
  ws3["A1"].value = text[:-1]
  wb.save(in_filename)


if __name__ == "__main__":
  if len(sys.argv) == 2:
    in_filename = sys.argv[1]
    samlokalisert(in_filename)  
  else:
    print("brukes som \"python3 script.py in_data.xlsx\"")

